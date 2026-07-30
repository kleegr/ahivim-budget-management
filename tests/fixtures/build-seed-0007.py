#!/usr/bin/env python3
"""Generate drizzle/0007_seed_calculation_strategies.sql from the workbook.

Seeds the Calculations-tab planning data as strategies linked to canonical
individuals. Uses the app's normalizePersonName (sorted parts) so an existing
individual is matched via ON CONFLICT (normalized_name) and never duplicated;
genuinely-new planning-only names are created (and can be merged later via the
aliases feature). Rates are NOT seeded here — they are read live from
program_rate_schedules. Idempotent: guarded with NOT EXISTS.
"""
import re, sys
from decimal import Decimal
import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else "tests/fixtures/workbook-source.xlsx"

def normalize(name):
    cleaned = re.sub(r"\s+", " ", re.sub(r"[^a-z, ]+", " ", (name or "").lower().replace("’","").replace("‘","").replace("'",""))).strip()
    if not cleaned: return ""
    parts = [p.strip() for seg in cleaned.split(",") for p in seg.split(" ") if p.strip()]
    return " ".join(sorted(parts))

def sq(s):  # SQL string literal
    return "'" + str(s).replace("'", "''") + "'"

PROG = {"ComHab":"COM_HAB","Respite":"RESPITE","SHCH":"SH_COM_HAB","SHR":"SH_RESPITE","DayHab":"DAY_HAB","SDH":"SUPP_GROUP_DAY_HAB"}
HRS = {"ComHab":7,"Respite":8,"SHCH":9,"SHR":10,"DayHab":11,"SDH":12}

wb = openpyxl.load_workbook(SRC, data_only=True)
ca = wb["Calculations"]

rows = []
for r in range(5, ca.max_row+1):
    raw = ca.cell(r,1).value
    if not raw: continue
    label_m = re.search(r"\s([12])\s*$", str(raw))
    label = label_m.group(1) if label_m else "1"
    base = re.sub(r"\s*\(.*?\)","", str(raw))
    base = re.sub(r"\s+[12]\s*$","", base).strip()
    norm = normalize(base)
    renew = ca.cell(r,2).value
    renew_iso = None
    if hasattr(renew, "isoformat"):
        y = renew.year
        renew_iso = renew.date().isoformat() if y >= 2000 else None
    yearly = sum((Decimal(str(ca.cell(r,HRS[k]).value or 0)) * Decimal("1") for k in HRS), Decimal(0))
    wb_month = Decimal(str(ca.cell(r,15).value or 0))
    # implied divisor from the workbook's own monthly figure
    yr = sum(Decimal(str(ca.cell(r,HRS[k]).value or 0)) for k in HRS)  # placeholder; real yearly uses rates
    divisor = 12
    # infer divisor from workbook yearly N / monthly O
    N = Decimal(str(ca.cell(r,14).value or 0)); O = Decimal(str(ca.cell(r,15).value or 0))
    if O and N:
        d = (N / O)
        divisor = int(d.to_integral_value()) if abs(d - d.to_integral_value()) < Decimal("0.05") else 12
    def dv(c):
        v = ca.cell(r,c).value
        return Decimal(str(v)) if v is not None else Decimal(0)
    cut1 = dv(3)/100; cut2 = dv(4)/100
    clock = dv(5); other = dv(6)
    after = ca.cell(r,18).value
    account = ca.cell(r,19).value
    hours = {k: Decimal(str(ca.cell(r,HRS[k]).value)) for k in HRS if ca.cell(r,HRS[k]).value not in (None,0)}
    rows.append(dict(display=base, norm=norm, label=label, renew=renew_iso, divisor=divisor,
                     cut1=cut1, cut2=cut2, clock=clock, other=other,
                     after=(Decimal(str(after)) if after is not None else None),
                     account=account, hours=hours, sort=int(label)))

out = []
out.append("-- Phase 5 (data): seed the Calculations-tab planning as strategies.")
out.append("-- Matched to existing individuals by the app's sorted-name normalization")
out.append("-- (ON CONFLICT (normalized_name) DO NOTHING) so no individual is duplicated;")
out.append("-- planning-only names are created and can be merged later via aliases.")
out.append("-- Idempotent (NOT EXISTS guards). Rates are read live from program_rate_schedules.")
out.append("")

# 1. individuals (distinct canonical)
seen=set()
for row in rows:
    if row["norm"] in seen: continue
    seen.add(row["norm"])
    out.append(f"INSERT INTO \"individuals\" (normalized_name, display_name, status) "
               f"VALUES ({sq(row['norm'])}, {sq(row['display'])}, 'active') "
               f"ON CONFLICT (normalized_name) DO NOTHING;--> statement-breakpoint")
out.append("")

# 2. strategies + lines
for row in rows:
    renew = "NULL" if not row["renew"] else f"DATE {sq(row['renew'])}"
    after = "NULL" if row["after"] is None else str(row["after"])
    account = "NULL" if not row["account"] else sq(row["account"])
    out.append(
        "INSERT INTO \"calculation_strategies\" "
        "(individual_id, label, renewal_date, month_divisor, cut1_percent, cut2_percent, "
        "clock_adjustment, other_adjustment, after_all, account, sort_order) "
        f"SELECT i.id, {sq(row['label'])}, {renew}, {row['divisor']}, {row['cut1']}, {row['cut2']}, "
        f"{row['clock']}, {row['other']}, {after}, {account}, {row['sort']} "
        f"FROM \"individuals\" i WHERE i.normalized_name = {sq(row['norm'])} "
        "AND NOT EXISTS (SELECT 1 FROM \"calculation_strategies\" s "
        f"WHERE s.individual_id = i.id AND s.label = {sq(row['label'])});--> statement-breakpoint")
    for k, h in row["hours"].items():
        code = PROG[k]
        out.append(
            "INSERT INTO \"calculation_strategy_lines\" (strategy_id, program_id, authorized_hours) "
            "SELECT s.id, p.id, " + str(h) + " "
            "FROM \"calculation_strategies\" s "
            "JOIN \"individuals\" i ON i.id = s.individual_id "
            f"JOIN \"programs\" p ON p.code = {sq(code)} "
            f"WHERE i.normalized_name = {sq(row['norm'])} AND s.label = {sq(row['label'])} "
            "AND NOT EXISTS (SELECT 1 FROM \"calculation_strategy_lines\" l "
            "WHERE l.strategy_id = s.id AND l.program_id = p.id);--> statement-breakpoint")
    out.append("")

sql = "\n".join(out).rstrip()
# drop the final breakpoint marker so the file ends cleanly
sql = sql.rsplit("--> statement-breakpoint",1)
sql = ("--> statement-breakpoint".join(sql[:-1]) + sql[-1]) if len(sql)>1 else sql[0]
open("drizzle/0007_seed_calculation_strategies.sql","w").write(sql + "\n")
print("wrote drizzle/0007_seed_calculation_strategies.sql")
print("strategies:", len(rows), " distinct individuals:", len(seen))
matched = 0
