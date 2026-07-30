#!/usr/bin/env python3
"""Regenerate tests/fixtures/workbook-parity.json from the source workbook.

This is the production-parity fixture: the Transactions and Calculations
workspaces must reproduce these numbers exactly (or show an explicit,
explained difference). Run: python3 tests/fixtures/build-workbook-parity.py PATH.xlsx
"""
import json, sys
from decimal import Decimal, ROUND_HALF_UP
import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else "workbook.xlsx"

def D(v):
    return Decimal(0) if v is None else Decimal(str(v))
def s(d):  # stringify decimal without exponent
    return format(d, "f")

wb = openpyxl.load_workbook(SRC, data_only=True)

# ---- Ahivim (actual billed history) ----
ah = wb["Ahivim"]
# row 1 = workbook totals, row 2 = headers, rows 3.. = data
COL = dict(payto=1, checkdate=2, checkno=3, code=4, hours=5, rate=6, gross=7,
           netpay=8, pbegin=9, pend=10, program=11, individual=12, employee=13,
           internal=16, netonce=19)
rows = []
for r in range(3, ah.max_row + 1):
    if ah.cell(r, COL["payto"]).value is None and ah.cell(r, COL["checkno"]).value is None and ah.cell(r, COL["gross"]).value is None:
        continue
    rows.append(r)

def sumcol(c):
    return sum((D(ah.cell(r, c).value) for r in rows), Decimal(0))

gross = sumcol(COL["gross"]); internal = sumcol(COL["internal"]); net_once = sumcol(COL["netonce"]); hours = sumcol(COL["hours"])
checks = {ah.cell(r, COL["checkno"]).value for r in rows if ah.cell(r, COL["checkno"]).value is not None}
inds = {ah.cell(r, COL["individual"]).value for r in rows if ah.cell(r, COL["individual"]).value}
emps = {ah.cell(r, COL["employee"]).value for r in rows if ah.cell(r, COL["employee"]).value}
progs = sorted({ah.cell(r, COL["program"]).value for r in rows if ah.cell(r, COL["program"]).value})
payees = sorted({ah.cell(r, COL["payto"]).value for r in rows if ah.cell(r, COL["payto"]).value})

# per-program filtered totals (gross/internal/hours + unique-check net within the filter)
def group_totals(keycol):
    out = {}
    for r in rows:
        k = ah.cell(r, keycol).value
        if k is None: continue
        k = str(k)
        g = out.setdefault(k, {"gross": Decimal(0), "internal": Decimal(0), "hours": Decimal(0),
                               "net_once": Decimal(0), "tx": 0, "checks": set()})
        g["gross"] += D(ah.cell(r, COL["gross"]).value)
        g["internal"] += D(ah.cell(r, COL["internal"]).value)
        g["hours"] += D(ah.cell(r, COL["hours"]).value)
        g["net_once"] += D(ah.cell(r, COL["netonce"]).value)
        g["tx"] += 1
        cn = ah.cell(r, COL["checkno"]).value
        if cn is not None: g["checks"].add(cn)
    return {k: {"gross": s(v["gross"]), "internal": s(v["internal"]), "hours": s(v["hours"]),
                "agencyAdditional": s(v["gross"] - v["internal"]), "netOnce": s(v["net_once"]),
                "transactions": v["tx"], "checks": len(v["checks"])} for k, v in out.items()}

ahivim = {
    "dataRows": len(rows),
    "totals": {
        "gross": s(gross), "internal": s(internal), "agencyAdditional": s(gross - internal),
        "netOnce": s(net_once), "hours": s(hours),
        "transactions": len(rows), "checks": len(checks),
        "individuals": len(inds), "employees": len(emps),
    },
    "workbookRow1": {"internalP": s(D(ah.cell(1,16).value)), "grossQ": s(D(ah.cell(1,17).value)),
                     "agencyAdditionalR": s(D(ah.cell(1,18).value)), "netS": s(D(ah.cell(1,19).value))},
    "programs": progs,
    "payees": payees,
    "byProgram": group_totals(COL["program"]),
    "byIndividual": group_totals(COL["individual"]),
    "byPayee": group_totals(COL["payto"]),
}

# ---- Calculations (planning) ----
ca = wb["Calculations"]
# row 1 headers, row 2 rates, rows 5.. = lines
RATE = {c: D(ca.cell(2, i).value) for c, i in [("ComHab",7),("Respite",8),("SHCH",9),("SHR",10),("DayHab",11),("SDH",12)]}
HRS = {"ComHab":7,"Respite":8,"SHCH":9,"SHR":10,"DayHab":11,"SDH":12}
calc_rows = []
for r in range(5, ca.max_row + 1):
    name = ca.cell(r, 1).value
    if not name: continue
    hrs = {k: D(ca.cell(r, c).value) for k, c in HRS.items()}
    yearly = sum((hrs[k] * RATE[k] for k in hrs), Decimal(0))
    wb_yearly = D(ca.cell(r, 14).value)
    wb_monthly = D(ca.cell(r, 15).value)
    # infer month divisor the workbook actually used
    divisor = (yearly / wb_monthly) if wb_monthly else Decimal(12)
    cut1 = D(ca.cell(r, 3).value); cut2 = D(ca.cell(r, 4).value)
    clock = D(ca.cell(r, 5).value); adj = D(ca.cell(r, 6).value)
    monthly = (yearly / Decimal(12))
    after1 = wb_monthly - (wb_monthly * cut1 / 100)
    after2 = after1 - (after1 * cut2 / 100)
    net = after2 + clock + adj
    calc_rows.append({
        "label": name,
        "renewal": ca.cell(r, 2).value.isoformat() if hasattr(ca.cell(r,2).value, "isoformat") else ca.cell(r,2).value,
        "cut1Pct": s(cut1), "cut2Pct": s(cut2), "clock": s(clock), "adjustment": s(adj),
        "hours": {k: s(v) for k, v in hrs.items()},
        "computedYearly": s(yearly),
        "workbookYearly": s(wb_yearly),
        "yearlyMatches": yearly == wb_yearly,
        "workbookMonthly": s(wb_monthly),
        "impliedMonthDivisor": s(divisor.quantize(Decimal("0.001"))),
        "workbookGrossNet": s(D(ca.cell(r, 16).value)),
        "computedGrossNet": s(after2.quantize(Decimal("0.000001"))),
        "workbookNet": s(D(ca.cell(r, 17).value)),
        "computedNet": s(net.quantize(Decimal("0.000001"))),
        "afterAll": s(D(ca.cell(r, 18).value)),
        "account": ca.cell(r, 19).value,
    })

canonical = {}
for cr in calc_rows:
    base = cr["label"].rstrip()
    # a trailing " 1"/" 2" strategy suffix maps to one canonical individual
    parts = base.rsplit(" ", 1)
    if len(parts) == 2 and parts[1].isdigit():
        canonical.setdefault(parts[0], []).append(cr["label"])
multi = {k: v for k, v in canonical.items() if len(v) > 1}

out = {
    "source": "Excellent Staffing 2025-2026.xlsx",
    "internalRates": {k: s(v) for k, v in RATE.items()},
    "ahivim": ahivim,
    "calculations": {"rates": {k: s(v) for k, v in RATE.items()},
                     "rows": calc_rows,
                     "multiStrategyIndividuals": multi},
}
with open("tests/fixtures/workbook-parity.json", "w") as f:
    json.dump(out, f, indent=2, default=str)
print("wrote tests/fixtures/workbook-parity.json")
print("Ahivim totals:", json.dumps(ahivim["totals"], indent=2))
print("Multi-strategy individuals:", json.dumps(multi, indent=2))
mism = [c["label"] for c in calc_rows if not c["yearlyMatches"]]
print("Yearly mismatches:", mism)
odddiv = [(c["label"], c["impliedMonthDivisor"]) for c in calc_rows if c["impliedMonthDivisor"] not in ("12.000",)]
print("Non-12 month divisors:", odddiv)
